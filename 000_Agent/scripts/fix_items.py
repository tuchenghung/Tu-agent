# -*- coding: utf-8 -*-
import openpyxl

FILE = r'D:\Dropbox\yushi\規劃中案件\20260608-南崁廠房高架地板工程\D預算報價\20260611南崁廠房高架地板工程.xlsx'

wb = openpyxl.load_workbook(FILE)
ws = wb.active

# 修正品名規格（B欄）
ws['B10'] = '鋁合金蜂巢板（開孔率17%沖孔加工費用)'
ws['B12'] = '踏階(2階)'
ws['B13'] = '施工完清潔'

wb.save(FILE)
wb.close()
print('saved')
